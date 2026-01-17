"""
Utilitaires pour l'export Excel et PDF
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from flask import current_app


def export_missions_to_excel(missions, start_date, end_date):
    """Exporte les missions en Excel"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Missions"

    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:N1')
    ws['A1'] = f'Planning des Missions - Du {start_date} au {end_date}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # En-têtes
    headers = [
        'Date', 'Heure', 'Type', 'Voyage', 'SST', 'Chauffeur',
        'Palettes', 'Numéro', 'Pays', 'Ramasse', 'Effectué',
        'Revenus', 'Coûts', 'Marge'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    for row_num, mission in enumerate(missions, 4):
        ws.cell(row=row_num, column=1, value=mission.date.isoformat() if mission.date else '')
        ws.cell(row=row_num, column=2, value=mission.heure)
        ws.cell(row=row_num, column=3, value=mission.type)
        ws.cell(row=row_num, column=4, value=mission.voyage)
        ws.cell(row=row_num, column=5, value=mission.sst)
        ws.cell(row=row_num, column=6, value=mission.chauffeur or '')
        ws.cell(row=row_num, column=7, value=mission.palettes)
        ws.cell(row=row_num, column=8, value=mission.numero or '')
        ws.cell(row=row_num, column=9, value=mission.pays)
        ws.cell(row=row_num, column=10, value='Oui' if mission.ramasse else 'Non')
        ws.cell(row=row_num, column=11, value='Oui' if mission.effectue else 'Non')
        ws.cell(row=row_num, column=12, value=mission.revenus)
        ws.cell(row=row_num, column=13, value=mission.couts)
        ws.cell(row=row_num, column=14, value=mission.marge)

        # Appliquer les bordures
        for col_num in range(1, 15):
            ws.cell(row=row_num, column=col_num).border = thin_border

    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 12, 'B': 8, 'C': 12, 'D': 12, 'E': 20,
        'F': 20, 'G': 10, 'H': 12, 'I': 12, 'J': 10,
        'K': 10, 'L': 12, 'M': 12, 'N': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Totaux
    if missions:
        last_row = len(missions) + 4
        ws.cell(row=last_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=last_row, column=7, value=f'=SUM(G4:G{last_row-1})')
        ws.cell(row=last_row, column=12, value=f'=SUM(L4:L{last_row-1})')
        ws.cell(row=last_row, column=13, value=f'=SUM(M4:M{last_row-1})')
        ws.cell(row=last_row, column=14, value=f'=SUM(N4:N{last_row-1})')

        for col_num in [1, 7, 12, 13, 14]:
            ws.cell(row=last_row, column=col_num).font = Font(bold=True)
            ws.cell(row=last_row, column=col_num).fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )

    # Sauvegarder
    export_dir = current_app.config['EXPORT_FOLDER']
    os.makedirs(export_dir, exist_ok=True)

    filename = f'missions_{start_date}_{end_date}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)

    wb.save(filepath)

    return filepath


def export_missions_to_pdf(missions, start_date, end_date):
    """Exporte les missions en PDF"""

    export_dir = current_app.config['EXPORT_FOLDER']
    os.makedirs(export_dir, exist_ok=True)

    filename = f'missions_{start_date}_{end_date}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    filepath = os.path.join(export_dir, filename)

    # Créer le document PDF en mode paysage
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=1  # Center
    )

    # Titre
    title = Paragraph(f'Planning des Missions<br/>Du {start_date} au {end_date}', title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))

    # Données du tableau
    data = [
        ['Date', 'Heure', 'Voyage', 'SST', 'Chauffeur', 'Pal.', 'Pays', 'Effectué', 'Revenus', 'Coûts', 'Marge']
    ]

    for mission in missions:
        data.append([
            mission.date.strftime('%d/%m/%Y') if mission.date else '',
            mission.heure,
            mission.voyage[:10],
            mission.sst[:15],
            (mission.chauffeur or '')[:15],
            str(mission.palettes),
            mission.pays[:8],
            'Oui' if mission.effectue else 'Non',
            f'{mission.revenus:.2f}€',
            f'{mission.couts:.2f}€',
            f'{mission.marge:.2f}€'
        ])

    # Totaux
    if missions:
        total_palettes = sum(m.palettes for m in missions)
        total_revenus = sum(m.revenus for m in missions)
        total_couts = sum(m.couts for m in missions)
        total_marge = sum(m.marge for m in missions)

        data.append([
            'TOTAL', '', '', '', '',
            str(total_palettes), '', '',
            f'{total_revenus:.2f}€',
            f'{total_couts:.2f}€',
            f'{total_marge:.2f}€'
        ])

    # Créer le tableau
    table = Table(data, repeatRows=1)

    # Style du tableau
    table_style = TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Corps du tableau
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),

        # Alternance de couleurs
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Générer le PDF
    doc.build(elements)

    return filepath


def export_chauffeurs_to_excel(chauffeurs):
    """Exporte la liste des chauffeurs en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chauffeurs"

    # En-têtes
    headers = ['Nom', 'SST', 'Actif', 'Informations']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Données
    for row_num, chauffeur in enumerate(chauffeurs, 2):
        ws.cell(row=row_num, column=1, value=chauffeur.nom)
        ws.cell(row=row_num, column=2, value=chauffeur.sst or '')
        ws.cell(row=row_num, column=3, value='Oui' if chauffeur.actif else 'Non')
        ws.cell(row=row_num, column=4, value=chauffeur.infos or '')

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40

    # Sauvegarder
    export_dir = current_app.config['EXPORT_FOLDER']
    os.makedirs(export_dir, exist_ok=True)

    filename = f'chauffeurs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(export_dir, filename)

    wb.save(filepath)

    return filepath
